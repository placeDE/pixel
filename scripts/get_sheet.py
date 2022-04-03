import openpyxl
from openpyxl import load_workbook
from PIL import Image

FILE_NAME = "master template.xlsx" # @param {type: "string"}

wb = load_workbook(FILE_NAME, data_only = True)
start_x = 0
start_y = 0

for sheet in wb:
  if(sheet.title == "Copy of BEEEEG"): continue
  TEMPLATE_WIDTH = sheet.max_column-1
  TEMPLATE_HEIGHT = sheet.max_row-1
  img = Image.new("RGBA", (TEMPLATE_WIDTH, TEMPLATE_HEIGHT))
  for row_idx in range(1, TEMPLATE_HEIGHT+2):
      for column_idx in range(1, TEMPLATE_WIDTH+2):
        cell = sheet.cell(row=row_idx, column=column_idx)
        hex_color = cell.fill.start_color.index
        if isinstance(hex_color, str) and len(hex_color) == 8:
            a, r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4, 6))
            rgba_color = (r, g, b, a)
        elif isinstance(hex_color, int):
            rgba_color = (0, 0, 0, 255) if hex_color else (255, 255, 255, 255)
        else:
            print(f"Bad color format: {hex_color}")
            continue
        img.putpixel((column_idx-2, row_idx-2), rgba_color)
  img.save("../images/"+sheet.title+".png")